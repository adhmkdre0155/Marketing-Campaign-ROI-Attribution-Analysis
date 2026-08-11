import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, ScatterChart, Reference, Series
from openpyxl.chart.marker import Marker
from openpyxl.utils import get_column_letter

NAVY = "1F3864"
GOLD = "B08D57"
LIGHT = "EAF1F8"
WHITE = "FFFFFF"

df = pd.read_csv("../data/campaign_performance_clean.csv", parse_dates=["Date"])
attr = pd.read_csv("../data/attribution_comparison.csv")

wb = Workbook()

# ---------------------------------------------------------------
# Sheet: Data (monthly aggregated, keeps workbook a manageable size)
# ---------------------------------------------------------------
monthly = df.groupby(["Month", "Channel"]).agg(
    Spend=("Spend", "sum"), Impressions=("Impressions", "sum"), Clicks=("Clicks", "sum"),
    Conversions=("Conversions", "sum"), Revenue=("Revenue", "sum")
).reset_index()
monthly["CAC"] = (monthly.Spend / monthly.Conversions).round(2)
monthly["ROAS"] = (monthly.Revenue / monthly.Spend).round(3)
monthly["ConversionRate"] = (monthly.Conversions / monthly.Clicks * 100).round(3)

ws_data = wb.active
ws_data.title = "Monthly_Data"
cols = ["Month", "Channel", "Spend", "Impressions", "Clicks", "Conversions", "Revenue", "CAC", "ROAS", "ConversionRate"]
ws_data.append(cols)
for c in ws_data[1]:
    c.font = Font(bold=True, color=WHITE, name="Arial")
    c.fill = PatternFill("solid", fgColor=NAVY)
for _, row in monthly.iterrows():
    ws_data.append([row[c] for c in cols])
for i in range(1, len(cols) + 1):
    ws_data.column_dimensions[get_column_letter(i)].width = 14
n_rows = len(monthly) + 1

def colref(name):
    return get_column_letter(cols.index(name) + 1)

# ---------------------------------------------------------------
# Sheet: Channel_Summary (formula-driven)
# ---------------------------------------------------------------
ws_c = wb.create_sheet("Channel_Summary")
channels = sorted(df.Channel.unique())
ws_c.append(["Channel", "TotalSpend", "TotalRevenue", "TotalConversions", "CAC", "ROAS", "SpendSharePct", "RevenueSharePct"])
for c in ws_c[1]:
    c.font = Font(bold=True, color=WHITE, name="Arial")
    c.fill = PatternFill("solid", fgColor=NAVY)

SP, REV, CONV = colref("Spend"), colref("Revenue"), colref("Conversions")
CH = colref("Channel")
for i, ch in enumerate(channels):
    r = i + 2
    ws_c.cell(row=r, column=1, value=ch)
    ws_c.cell(row=r, column=2, value=f'=ROUND(SUMIF(Monthly_Data!${CH}$2:${CH}${n_rows},A{r},Monthly_Data!${SP}$2:${SP}${n_rows}),0)')
    ws_c.cell(row=r, column=3, value=f'=ROUND(SUMIF(Monthly_Data!${CH}$2:${CH}${n_rows},A{r},Monthly_Data!${REV}$2:${REV}${n_rows}),0)')
    ws_c.cell(row=r, column=4, value=f'=SUMIF(Monthly_Data!${CH}$2:${CH}${n_rows},A{r},Monthly_Data!${CONV}$2:${CONV}${n_rows})')
    ws_c.cell(row=r, column=5, value=f'=ROUND(B{r}/D{r},2)')
    ws_c.cell(row=r, column=6, value=f'=ROUND(C{r}/B{r},2)')
    ws_c.cell(row=r, column=7, value=f'=ROUND(B{r}/SUM($B$2:$B${1+len(channels)})*100,1)')
    ws_c.cell(row=r, column=8, value=f'=ROUND(C{r}/SUM($C$2:$C${1+len(channels)})*100,1)')
last_c_row = len(channels) + 1
for i in range(1, 9):
    ws_c.column_dimensions[get_column_letter(i)].width = 17

# ---------------------------------------------------------------
# Sheet: Attribution_Comparison
# ---------------------------------------------------------------
ws_a = wb.create_sheet("Attribution_Comparison")
ws_a.append(["Channel", "FirstTouchRevenue", "LastTouchRevenue", "LinearRevenue", "FirstVsLast_PctDiff"])
for c in ws_a[1]:
    c.font = Font(bold=True, color=WHITE, name="Arial")
    c.fill = PatternFill("solid", fgColor=NAVY)
for i, row in attr.iterrows():
    r = i + 2
    ws_a.cell(row=r, column=1, value=row["Channel"])
    ws_a.cell(row=r, column=2, value=round(row["FirstTouchRevenue"], 0))
    ws_a.cell(row=r, column=3, value=round(row["LastTouchRevenue"], 0))
    ws_a.cell(row=r, column=4, value=round(row["LinearRevenue"], 0))
    ws_a.cell(row=r, column=5, value=round(row["FirstVsLast_PctDiff"], 1))
last_a_row = len(attr) + 1
for i in range(1, 6):
    ws_a.column_dimensions[get_column_letter(i)].width = 20

# ---------------------------------------------------------------
# Sheet: Reallocation_Scenario
# ---------------------------------------------------------------
ws_s = wb.create_sheet("Reallocation_Scenario")
ws_s["A1"] = "SCENARIO: Shift Budget from Paid Social (saturated) to Email (headroom)"
ws_s["A1"].font = Font(bold=True, size=13, color=NAVY, name="Arial")
labels = ["Current Blended ROAS", "Budget Shift Amount (€)", "Paid Social Marginal ROAS (late period)",
          "Email Conservative Incremental ROAS", "Paid Social Revenue Lost (€)",
          "Email Revenue Gained (€)", "New Blended ROAS", "Projected ROAS Lift (%)"]
for i, lbl in enumerate(labels):
    ws_s.cell(row=3 + i, column=1, value=lbl).font = Font(bold=True, name="Arial")

total_spend_formula = f"SUM(Channel_Summary!B2:B{last_c_row})"
total_revenue_formula = f"SUM(Channel_Summary!C2:C{last_c_row})"
ws_s["B3"] = f"=ROUND({total_revenue_formula}/{total_spend_formula},2)"
ws_s["B4"] = 75000
ws_s["B5"] = 1.94
ws_s["B6"] = 50.0
ws_s["B7"] = "=ROUND(B4*B5,0)"
ws_s["B8"] = "=ROUND(B4*B6,0)"
ws_s["B9"] = f"=ROUND(({total_revenue_formula}-B7+B8)/{total_spend_formula},2)"
ws_s["B10"] = "=ROUND((B9-B3)/B3*100,1)"
for r in range(3, 11):
    ws_s.cell(row=r, column=2).font = Font(size=12, color=GOLD, bold=True, name="Arial")
ws_s["B10"].font = Font(size=15, color="E34948", bold=True, name="Arial")
ws_s["A12"] = "Note: Email's incremental ROAS is deliberately capped well below its observed ~148x average (a conservative ~1/3 discount) to account for list-size and deliverability limits not captured in this dataset."
ws_s["A12"].font = Font(italic=True, size=9, color="666666", name="Arial")
ws_s.column_dimensions["A"].width = 42
ws_s.column_dimensions["B"].width = 20

# ---------------------------------------------------------------
# Sheet: Dashboard
# ---------------------------------------------------------------
ws_d = wb.create_sheet("Dashboard", 0)
ws_d.sheet_view.showGridLines = False
ws_d.merge_cells("B2:K2")
ws_d["B2"] = "MARKETING CAMPAIGN ROI & ATTRIBUTION DASHBOARD"
ws_d["B2"].font = Font(bold=True, size=18, color=NAVY, name="Arial")
ws_d.merge_cells("B3:K3")
ws_d["B3"] = "SEO · Paid Social · Paid Search · Email — 2024-2025"
ws_d["B3"].font = Font(italic=True, size=12, color=GOLD, name="Arial")

def kpi_card(ws, col, label, formula, fmt="#,##0"):
    col_letter = get_column_letter(col)
    ws.merge_cells(f"{col_letter}5:{get_column_letter(col+1)}5")
    ws[f"{col_letter}5"] = label
    ws[f"{col_letter}5"].font = Font(bold=True, color=WHITE, size=9, name="Arial")
    ws[f"{col_letter}5"].fill = PatternFill("solid", fgColor=NAVY)
    ws[f"{col_letter}5"].alignment = Alignment(horizontal="center", wrap_text=True)
    ws.merge_cells(f"{col_letter}6:{get_column_letter(col+1)}7")
    cell = ws[f"{col_letter}6"]
    cell.value = formula
    cell.font = Font(bold=True, size=16, color=GOLD, name="Arial")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = fmt
    for rr in (5, 6, 7):
        for cc in (col, col + 1):
            ws.cell(row=rr, column=cc).border = Border(*(Side(style="thin", color="CCCCCC"),) * 4)

kpi_card(ws_d, 2, "BLENDED ROAS", "=Reallocation_Scenario!B3", '0.00"x"')
kpi_card(ws_d, 4, "PAID SOCIAL SPEND SHARE", "=Channel_Summary!G4", '0.0"%"')
kpi_card(ws_d, 6, "PAID SOCIAL REVENUE SHARE", "=Channel_Summary!H4", '0.0"%"')
kpi_card(ws_d, 8, "PROJECTED ROAS LIFT", "=Reallocation_Scenario!B10", '0.0"%"')
kpi_card(ws_d, 10, "TOP CHANNEL: FIRST-TOUCH", f'=INDEX(Attribution_Comparison!A2:A{last_a_row},MATCH(MAX(Attribution_Comparison!B2:B{last_a_row}),Attribution_Comparison!B2:B{last_a_row},0))', "@")

ws_d.row_dimensions[6].height = 20
ws_d.row_dimensions[7].height = 20

bar = BarChart()
bar.title = "ROAS by Channel"
bar.y_axis.title = "ROAS (x)"
data = Reference(ws_c, min_col=6, min_row=1, max_row=last_c_row)
cats = Reference(ws_c, min_col=1, min_row=2, max_row=last_c_row)
bar.add_data(data, titles_from_data=True)
bar.set_categories(cats)
bar.width, bar.height = 16, 9
ws_d.add_chart(bar, "B10")

bar2 = BarChart()
bar2.type = "bar"
bar2.title = "First-Touch vs. Last-Touch Revenue Credit"
bar2.y_axis.title = "Revenue (€)"
data2 = Reference(ws_a, min_col=2, max_col=3, min_row=1, max_row=last_a_row)
cats2 = Reference(ws_a, min_col=1, min_row=2, max_row=last_a_row)
bar2.add_data(data2, titles_from_data=True)
bar2.set_categories(cats2)
bar2.width, bar2.height = 16, 9
ws_d.add_chart(bar2, "B29")

for i in range(1, 12):
    ws_d.column_dimensions[get_column_letter(i)].width = 15
ws_d.page_setup.orientation = "landscape"
ws_d.page_setup.fitToWidth = 1
ws_d.page_setup.fitToHeight = 0
ws_d.sheet_properties.pageSetUpPr.fitToPage = True

wb.save("Marketing_ROI_Dashboard.xlsx")
print("saved")
